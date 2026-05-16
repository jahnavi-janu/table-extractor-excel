import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """Handles creation of Excel files from extracted table data"""
    
    def __init__(self):
        """Initialize Excel writer"""
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_excel(self, tables: List[Dict[str, Any]], single_sheet: bool = True) -> bytes:
        """
        Create Excel workbook from extracted tables
        
        Args:
            tables: List of table dictionaries with 'title' and 'data' keys
            
        Returns:
            Excel file as bytes
        """
        if single_sheet:
            return self._create_single_sheet_excel(tables)
        return self._create_multi_sheet_excel(tables)

    def _create_single_sheet_excel(self, tables: List[Dict[str, Any]]) -> bytes:
        """Create workbook where all tables are written into one sheet."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Extracted Tables"

        current_row = 1
        max_col_widths = {}
        # Detect if all tables have identical headers (same order and names)
        dfs = [t.get("data") for t in tables if t.get("data") is not None]
        same_headers = False
        if dfs:
            try:
                first_cols = list(dfs[0].columns)
                same_headers = all(list(df.columns) == first_cols for df in dfs)
            except Exception:
                same_headers = False

        if same_headers:
            # Write a single header row
            for col_idx, header in enumerate(first_cols, 1):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = self.border
                max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), len(str(header)))

            worksheet.row_dimensions[current_row].height = 30
            current_row += 1

            # Append all rows from all tables under the same header
            for df in dfs:
                for row in df.values:
                    for col_idx, value in enumerate(row, 1):
                        cell = worksheet.cell(row=current_row, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = self.border
                        max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), len(str(value)))
                    current_row += 1

        else:
            # Fallback: write each table with its own header and title
            for idx, table_info in enumerate(tables, 1):
                df = table_info.get('data')
                if df is None:
                    continue

                title = table_info.get('title', f'Table {idx}')
                title_cell = worksheet.cell(row=current_row, column=1)
                title_cell.value = f"Table {idx}: {title}"
                title_cell.font = Font(bold=True, size=12)
                current_row += 1

                for col_idx, header in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx)
                    cell.value = header
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = self.border
                    max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), len(str(header)))

                worksheet.row_dimensions[current_row].height = 30
                current_row += 1

                for row in df.values:
                    for col_idx, value in enumerate(row, 1):
                        cell = worksheet.cell(row=current_row, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = self.border
                        max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), len(str(value)))
                    current_row += 1

                description = table_info.get('description', '')
                if description:
                    desc_cell = worksheet.cell(row=current_row, column=1)
                    desc_cell.value = f"Description: {description}"
                    desc_cell.font = Font(italic=True, size=9, color="666666")
                    current_row += 1

                current_row += 1

        for col_idx, max_length in max_col_widths.items():
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()

    def _create_multi_sheet_excel(self, tables: List[Dict[str, Any]]) -> bytes:
        """Create workbook with one table per worksheet."""
        workbook = Workbook()
        workbook.remove(workbook.active)
        used_sheet_names = set()

        for idx, table_info in enumerate(tables, 1):
            df = table_info.get("data")
            if df is None:
                continue

            title = table_info.get("title", f"Table {idx}")
            base_name = self._sanitize_sheet_name(str(title)[:31])
            sheet_name = self._unique_sheet_name(base_name, used_sheet_names)
            used_sheet_names.add(sheet_name)

            worksheet = workbook.create_sheet(title=sheet_name)

            for col_idx, header in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = self.border

            worksheet.row_dimensions[1].height = 30

            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = self.border

            for col_idx, column in enumerate(df.columns, 1):
                column_values = [str(column)] + [str(value) for value in df[column].tolist()]
                max_length = max(len(value) for value in column_values)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

            description = table_info.get("description", "")
            if description:
                desc_row = len(df) + 3
                desc_cell = worksheet.cell(row=desc_row, column=1)
                desc_cell.value = f"Description: {description}"
                desc_cell.font = Font(italic=True, size=9, color="666666")

        if not workbook.sheetnames:
            workbook.create_sheet(title="Extracted Tables")

        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """
        Sanitize sheet name for Excel
        
        Args:
            name: Original sheet name
            
        Returns:
            Sanitized sheet name
        """
        # Remove invalid characters
        invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            name = name.replace(char, '')
        
        # Replace spaces with underscores if name is too long
        name = name.strip()
        if not name:
            name = "Sheet"
        
        return name[:31]

    def _unique_sheet_name(self, base_name: str, used_names: set) -> str:
        """Create a unique sheet name within Excel's 31-char limit."""
        if base_name not in used_names:
            return base_name

        suffix = 2
        while True:
            suffix_text = f"_{suffix}"
            candidate = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
            if candidate not in used_names:
                return candidate
            suffix += 1
